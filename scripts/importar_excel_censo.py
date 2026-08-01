#!/usr/bin/env python3
"""Importa un Excel de censo externo con verificación Nexus y seguridad.

Uso típico:

  NEXUS_SCRIPT_EMAIL=admin@refugio.app NEXUS_SCRIPT_PASSWORD=... \\
    python3 scripts/importar_excel_censo.py \\
      --archivo /tmp/CE_Andres_Bello.xlsx --centro-id centro-32 --con-nexus --dry-run

  NEXUS_SCRIPT_EMAIL=admin@refugio.app NEXUS_SCRIPT_PASSWORD=... \\
    python3 scripts/importar_excel_censo.py \\
      --archivo /tmp/CE_Andres_Bello.xlsx --centro-id centro-32 --con-nexus --aplicar
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from collections import Counter
from concurrent.futures import FIRST_COMPLETED, Future, ThreadPoolExecutor, wait
from dataclasses import dataclass
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
ENV = ROOT / ".env"
DEFAULT_URL = "https://xzwifkckkakldnzkdeby.supabase.co"
DEFAULT_GATEWAY = "https://nexus.m0n1t0r-d3-3v3nt0s.net"
DEFAULT_CONCURRENCY = 5
DEFAULT_RATE = 2.5
DEFAULT_CIRCUIT_BREAKER = 5
DEFAULT_NEXUS_TIMEOUT = 20.0
PROGRESS_INTERVAL = 10.0
PAGE_SIZE = 1000
SIN_CEDULA = {
    "",
    "S/C",
    "SC",
    "S/N",
    "SN",
    "S/D",
    "SD",
    "SIN CEDULA",
    "SIN CÉDULA",
    "SIN DOCUMENTO",
    "SIN DOC",
    "N/A",
    "NA",
    "-",
    "NO",
    "NO TIENE",
    "NO POSEE",
    "NO POSSÉ",
    "NO APORTO",
    "NO APORTÓ",
    "NOPOSEE",
    "NOTIENE",
}
LETRAS_NEXUS = {"V", "E"}
SIN_CEDULA_KEYS: set[str] | None = None

ALIASES_NOMBRE: dict[str, str] = {
    "delgado chalbaud": "centro-33",
    "delgado chalboud": "centro-33",
    "perez bonalde": "centro-34",
    "pérez bonalde": "centro-34",
    "bonalde": "centro-34",
    "lossada": "centro-11",
    "losada": "centro-11",
    "mama rosa": "centro-36",
    "mamá rosa": "centro-36",
}

PREFIJOS = re.compile(
    r"^(u\.?\s*e\.?\s*n\.?\s*b\.?|u\.?\s*e\.?\s*n\.?|u\.?\s*e\.?\s*d\.?|u\.?\s*e\.?\s*e\.?|"
    r"u\.?\s*e\.?\s*|e\.?\s*n\.?\s*b\.?|e\.?\s*b\.?\s*e\.?|e\.?\s*b\.?\s*n\.?|"
    r"e\.?\s*t\.?\s*i\.?|c\.?\s*e\.?\s*i\.?\s*n\.?|c\.?\s*e\.?\s*i\.?|c\.?\s*e\.?\s*n\.?|"
    r"c\.?\s*e\.?\s*|complejo\s+educativo|liceo|escuela(\s+integral\s+basica)?|"
    r"unidad\s+educativa|refugio(\s+para)?|universidad|polideportivo|estadio|"
    r"fundacion|campamento|centro\s+de\s+educacion\s+inicial)\s+",
    re.I,
)


@dataclass
class CentroApp:
    id: str
    nombre: str
    activo: bool


def strip_accents(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto or "") if unicodedata.category(c) != "Mn"
    )


def key(texto: str) -> str:
    texto = strip_accents(str(texto or "").lower())
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _sin_cedula_keys() -> set[str]:
    global SIN_CEDULA_KEYS
    if SIN_CEDULA_KEYS is None:
        SIN_CEDULA_KEYS = {key(m) for m in SIN_CEDULA}
    return SIN_CEDULA_KEYS


def es_marcador_sin_cedula(valor: str) -> bool:
    """True si el texto declara ausencia de documento (no posee, s/c, etc.)."""
    return key(valor) in _sin_cedula_keys()


def normalizar_nombre_centro(texto: str) -> str:
    texto = strip_accents((texto or "").lower()).replace("ü", "u")
    texto = PREFIJOS.sub("", texto)
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def tokens(texto: str) -> set[str]:
    return {t for t in normalizar_nombre_centro(texto).split() if len(t) > 1}


def score_nombres(a: str, b: str) -> float:
    na, nb = normalizar_nombre_centro(a), normalizar_nombre_centro(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return 0.92
    ta, tb = tokens(a), tokens(b)
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    union = len(ta | tb)
    jaccard = inter / union
    nucleo_a = na.split()[-1] if na.split() else ""
    nucleo_b = nb.split()[-1] if nb.split() else ""
    if nucleo_a and nucleo_a == nucleo_b and inter >= 1:
        jaccard = max(jaccard, 0.75)
    return jaccard


def cargar_env() -> tuple[str, str, str]:
    valores: dict[str, str] = {}
    claves = (
        "VITE_SUPABASE_URL",
        "VITE_SUPABASE_ANON_KEY",
        "VITE_NEXUS_GATEWAY_URL",
        "NEXUS_SCRIPT_EMAIL",
        "NEXUS_SCRIPT_PASSWORD",
    )
    if ENV.exists():
        for linea in ENV.read_text(encoding="utf-8").splitlines():
            linea = linea.strip()
            for clave in claves:
                if linea.startswith(f"{clave}="):
                    valores[clave] = linea.split("=", 1)[1].strip().strip('"').strip("'")
    # Credenciales de script: .env basta; env del shell gana si ya están seteadas.
    for clave in ("NEXUS_SCRIPT_EMAIL", "NEXUS_SCRIPT_PASSWORD"):
        if valores.get(clave) and not os.environ.get(clave):
            os.environ[clave] = valores[clave]
    url = os.environ.get("VITE_SUPABASE_URL", valores.get("VITE_SUPABASE_URL", DEFAULT_URL))
    anon = os.environ.get("VITE_SUPABASE_ANON_KEY", valores.get("VITE_SUPABASE_ANON_KEY", ""))
    gateway = os.environ.get(
        "VITE_NEXUS_GATEWAY_URL",
        valores.get("VITE_NEXUS_GATEWAY_URL", DEFAULT_GATEWAY),
    )
    if not anon:
        raise SystemExit("Falta VITE_SUPABASE_ANON_KEY (.env o variable de entorno)")
    return url.rstrip("/"), anon, gateway.rstrip("/")


def autenticar(url: str, anon_key: str) -> str:
    email = os.environ.get("NEXUS_SCRIPT_EMAIL")
    password = os.environ.get("NEXUS_SCRIPT_PASSWORD")
    if not email or not password:
        raise SystemExit(
            "Faltan NEXUS_SCRIPT_EMAIL / NEXUS_SCRIPT_PASSWORD (.env o variables de entorno)"
        )
    req = urllib.request.Request(
        f"{url}/auth/v1/token?grant_type=password",
        data=json.dumps({"email": email, "password": password}).encode("utf-8"),
        headers={"apikey": anon_key, "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detalle = exc.read().decode("utf-8", errors="replace")
        raise SystemExit(f"No se pudo autenticar ({exc.code}): {detalle}") from exc
    token = body.get("access_token")
    if not token:
        raise SystemExit(f"Respuesta de login sin access_token: {body}")
    return token


def rpc(url: str, anon_key: str, jwt: str, fn: str, payload: dict[str, Any]) -> object:
    req = urllib.request.Request(
        f"{url}/rest/v1/rpc/{fn}",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "apikey": anon_key,
            "Authorization": f"Bearer {jwt}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as exc:
        detalle = exc.read().decode("utf-8", errors="replace")
        raise SystemExit(f"RPC {fn} falló HTTP {exc.code}: {detalle}") from exc


def rest_get_paginado(url: str, anon_key: str, jwt: str, path: str) -> list[dict[str, Any]]:
    """Lee todos los resultados PostgREST respetando límite por página."""
    acumulado: list[dict[str, Any]] = []
    desde = 0
    while True:
        hasta = desde + PAGE_SIZE - 1
        req = urllib.request.Request(
            f"{url}/rest/v1/{path}",
            headers={
                "apikey": anon_key,
                "Authorization": f"Bearer {jwt}",
                "Range-Unit": "items",
                "Range": f"{desde}-{hasta}",
            },
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            pagina = json.loads(resp.read().decode("utf-8"))
        if not isinstance(pagina, list):
            raise SystemExit(f"Respuesta REST inesperada para {path}")
        acumulado.extend(pagina)
        if len(pagina) < PAGE_SIZE:
            return acumulado
        desde += PAGE_SIZE


def cargar_cache_nexus(
    url: str,
    anon_key: str,
    jwt: str,
) -> dict[tuple[str, str], dict[str, Any]]:
    """Carga verificaciones persistentes. Una clave presente jamás toca Nexus."""
    filas = rest_get_paginado(
        url,
        anon_key,
        jwt,
        "nexus_consultas?select=letra,cedula,data,actualizado_ts&order=letra,cedula",
    )
    cache: dict[tuple[str, str], dict[str, Any]] = {}
    for fila in filas:
        letra = texto(fila.get("letra")).upper()
        cedula = texto(fila.get("cedula"))
        data = fila.get("data")
        if letra in LETRAS_NEXUS and cedula and isinstance(data, dict) and data.get("ok") is not False:
            cache[(letra, cedula)] = data
    return cache


def guardar_cache_nexus(
    url: str,
    anon_key: str,
    jwt: str,
    fichas: dict[tuple[str, str], dict[str, Any]],
) -> str:
    """Persiste fichas nuevas en lotes; próximas corridas no consultan Nexus.

    Reautentica si el JWT expiró (corridas Nexus largas > ~1h). Devuelve el JWT
    vigente tras el guardado.
    """
    if not fichas:
        return jwt
    filas = [
        {
            "letra": letra,
            "cedula": cedula,
            "data": data,
            "actualizado_ts": int(time.time() * 1000),
            "actualizado_por": "script:importar_excel_censo",
        }
        for (letra, cedula), data in fichas.items()
    ]
    token = jwt
    for inicio in range(0, len(filas), 200):
        lote = filas[inicio : inicio + 200]
        for intento in range(2):
            req = urllib.request.Request(
                f"{url}/rest/v1/nexus_consultas?on_conflict=letra,cedula",
                data=json.dumps(lote).encode("utf-8"),
                headers={
                    "apikey": anon_key,
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json",
                    "Prefer": "resolution=merge-duplicates,return=minimal",
                },
                method="POST",
            )
            try:
                with urllib.request.urlopen(req, timeout=60):
                    break
            except urllib.error.HTTPError as exc:
                if exc.code == 401 and intento == 0:
                    print("Nexus: JWT expirado al guardar caché · reautenticando…", file=sys.stderr, flush=True)
                    token = autenticar(url, anon_key)
                    continue
                raise
    return token


def listar_centros(url: str, anon_key: str, jwt: str) -> list[CentroApp]:
    req = urllib.request.Request(
        f"{url}/rest/v1/centros?select=id,data,deleted&limit=500",
        headers={"apikey": anon_key, "Authorization": f"Bearer {jwt}"},
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        rows = json.loads(resp.read().decode("utf-8"))
    return [
        CentroApp(
            id=row["id"],
            nombre=((row.get("data") or {}).get("nombre") or row["id"]).strip(),
            activo=not bool(row.get("deleted")),
        )
        for row in rows
    ]


def consultar_nexus(
    gateway: str,
    jwt: str,
    letra: str,
    cedula: str,
    timeout: float,
) -> dict[str, Any]:
    req = urllib.request.Request(
        f"{gateway}/v1/person/search/external/full/{letra}/{cedula}/censo",
        data=b"{}",
        headers={"Authorization": f"Bearer {jwt}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return {"status": resp.status, "body": json.loads(resp.read().decode("utf-8"))}
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        try:
            body = json.loads(raw)
        except json.JSONDecodeError:
            body = {"error": raw}
        return {"status": exc.code, "body": body}
    except urllib.error.URLError as exc:
        return {"status": None, "body": {"error": str(exc.reason)}}
    except TimeoutError as exc:
        return {"status": None, "body": {"error": f"timeout: {exc}"}}
    except OSError as exc:
        return {"status": None, "body": {"error": str(exc)}}


def consultar_nexus_concurrente(
    gateway: str,
    jwt: str,
    pendientes: list[tuple[str, str]],
    concurrency: int,
    rate: float,
    circuit_breaker: int,
    timeout: float,
    flush_fichas: Any | None = None,
) -> tuple[
    dict[tuple[str, str], dict[str, Any]],
    list[dict[str, Any]],
    int,
]:
    """Consulta claves únicas con límite calibrado y corte ante caída.

    Con concurrencia >1, envía una solicitud cada 1/N segundos y mantiene como
    máximo N en vuelo. Concurrencia 1 conserva modo secuencial con ``rate``.
    ``flush_fichas`` opcional se llama cada 100 fichas OK (y al final) para
    persistir caché incremental en corridas largas.
    """
    fichas: dict[tuple[str, str], dict[str, Any]] = {}
    pendientes_flush: dict[tuple[str, str], dict[str, Any]] = {}
    errores: list[dict[str, Any]] = []
    omitidas = 0
    fallos_consecutivos = 0
    procesadas = 0
    total = len(pendientes)
    ultimo_progreso = time.monotonic()

    def flush_si_toca(forzar: bool = False) -> None:
        if flush_fichas is None or not pendientes_flush:
            return
        if not forzar and len(pendientes_flush) < 100:
            return
        lote = dict(pendientes_flush)
        pendientes_flush.clear()
        flush_fichas(lote)
        print(
            f"Nexus: caché persistida · +{len(lote)} fichas (acum {len(fichas)})",
            file=sys.stderr,
            flush=True,
        )

    def procesar(clave: tuple[str, str]) -> tuple[tuple[str, str], dict[str, Any]]:
        letra, cedula = clave
        return clave, consultar_nexus(gateway, jwt, letra, cedula, timeout)

    def registrar(
        clave: tuple[str, str],
        resultado: dict[str, Any],
    ) -> bool:
        nonlocal fallos_consecutivos, procesadas, ultimo_progreso
        procesadas += 1
        status = resultado.get("status")
        body = resultado.get("body")
        if status == 200 and isinstance(body, dict) and body.get("ok") is not False:
            fichas[clave] = body
            pendientes_flush[clave] = body
            fallos_consecutivos = 0
            flush_si_toca()
        else:
            if status == 404 or (status == 200 and isinstance(body, dict) and body.get("ok") is False):
                fallos_consecutivos = 0
            else:
                fallos_consecutivos += 1
            errores.append(
                {
                    "tipo_doc": clave[0],
                    "documento": clave[1],
                    "status": status,
                    "detalle": body,
                }
            )
        ahora = time.monotonic()
        if procesadas == total or procesadas % 10 == 0 or ahora - ultimo_progreso >= PROGRESS_INTERVAL:
            print(
                f"Nexus: {procesadas}/{total} procesadas · "
                f"{len(fichas)} verificadas · {len(errores)} errores",
                file=sys.stderr,
                flush=True,
            )
            ultimo_progreso = ahora
        return fallos_consecutivos >= circuit_breaker

    if concurrency <= 1:
        for indice, clave in enumerate(pendientes):
            if indice > 0:
                time.sleep(rate)
            _, resultado = procesar(clave)
            if registrar(clave, resultado):
                omitidas = len(pendientes) - indice - 1
                break
        flush_si_toca(forzar=True)
        return fichas, errores, omitidas

    intervalo = 1.0 / concurrency
    proximas = iter(pendientes)
    en_vuelo: dict[Future[tuple[tuple[str, str], dict[str, Any]]], tuple[str, str]] = {}
    abortado = False
    ultimo_envio: float | None = None

    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        while not abortado:
            while len(en_vuelo) < concurrency:
                try:
                    clave = next(proximas)
                except StopIteration:
                    break
                if ultimo_envio is not None:
                    espera = intervalo - (time.monotonic() - ultimo_envio)
                    if espera > 0:
                        time.sleep(espera)
                futuro = executor.submit(procesar, clave)
                ultimo_envio = time.monotonic()
                en_vuelo[futuro] = clave

            if not en_vuelo:
                break

            terminados, _ = wait(en_vuelo, timeout=1, return_when=FIRST_COMPLETED)
            if not terminados:
                ahora = time.monotonic()
                if ahora - ultimo_progreso >= PROGRESS_INTERVAL:
                    print(
                        f"Nexus: {procesadas}/{total} procesadas · "
                        f"{len(en_vuelo)} en vuelo (timeout {timeout:g}s)",
                        file=sys.stderr,
                        flush=True,
                    )
                    ultimo_progreso = ahora
                continue
            for futuro in terminados:
                clave = en_vuelo.pop(futuro)
                try:
                    _, resultado = futuro.result()
                except Exception as exc:  # noqa: BLE001
                    resultado = {"status": None, "body": {"error": str(exc)}}
                if registrar(clave, resultado):
                    abortado = True
                    break

        if abortado:
            omitidas = sum(1 for _ in proximas)
            for futuro, clave in list(en_vuelo.items()):
                try:
                    _, resultado = futuro.result()
                except Exception as exc:  # noqa: BLE001
                    resultado = {"status": None, "body": {"error": str(exc)}}
                registrar(clave, resultado)

    flush_si_toca(forzar=True)
    return fichas, errores, omitidas


def resolver_centro(nombre_raw: str, centros: list[CentroApp], forzado: str | None) -> tuple[str | None, str, str]:
    if forzado:
        centro = next((c for c in centros if c.id == forzado), None)
        if centro is None:
            return None, nombre_raw, "inexistente"
        if not centro.activo:
            return None, nombre_raw, "inactivo"
        return forzado, nombre_raw, "forzado"
    raw = (nombre_raw or "").strip()
    if not raw:
        return None, "", ""
    nombre_key = normalizar_nombre_centro(raw)
    if nombre_key in ALIASES_NOMBRE:
        centro_id = ALIASES_NOMBRE[nombre_key]
        centro = next((c for c in centros if c.id == centro_id), None)
        if centro is None:
            return None, raw, "inexistente"
        if not centro.activo:
            return None, raw, "inactivo"
        return centro_id, raw, "alias"
    mejor: CentroApp | None = None
    mejor_score = 0.0
    for centro in centros:
        puntaje = score_nombres(raw, centro.nombre)
        if puntaje > mejor_score:
            mejor_score = puntaje
            mejor = centro
    if mejor and not mejor.activo and mejor_score >= 0.75:
        return None, raw, "inactivo"
    if mejor and mejor_score >= 0.99:
        return mejor.id, raw, "exacto"
    if mejor and mejor_score >= 0.75:
        return mejor.id, raw, "fuzzy"
    return None, raw, ""


def _es_fila_header(celdas: list[str]) -> bool:
    """Detecta fila de encabezados reales (salta título mergeado arriba)."""
    unidos = " ".join(key(c) for c in celdas if c)
    marcadores = (
        "primer nombre",
        "nombre completo",
        "documento",
        "cedula",
        "cédula",
        "apellido",
        "campamento",
    )
    hits = sum(1 for m in marcadores if m in unidos)
    return hits >= 2


def leer_filas(path: Path) -> tuple[list[dict[str, Any]], str, list[str]]:
    sufijo = path.suffix.lower()
    if sufijo in {".xlsx", ".xlsm"}:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise SystemExit("Para .xlsx instale openpyxl: pip install openpyxl") from exc
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = None
        rows: list[tuple[Any, ...]] = []
        header_idx = -1
        for candidata in wb.worksheets:
            candidatas = list(candidata.iter_rows(min_row=1, max_row=20, values_only=True))
            for i, row in enumerate(candidatas):
                celdas = [str(h or "").strip() for h in row]
                if _es_fila_header(celdas):
                    ws = candidata
                    header_idx = i
                    break
            if ws is not None:
                break
        if ws is None:
            raise SystemExit("No se encontró una hoja con encabezados de censo")
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h or "").strip() for h in rows[header_idx]]
        salida: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            item = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[i] if i < len(row) else None
            if any(v not in (None, "") for v in item.values()):
                salida.append(item)
        return salida, ws.title, headers

    text = path.read_text(encoding="utf-8-sig")
    dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    headers = [str(h or "").strip() for h in (reader.fieldnames or [])]
    return [{k.strip(): v for k, v in row.items() if k} for row in reader], path.name, headers


def pick(row: dict[str, Any], *aliases: str) -> Any:
    """Busca valor por alias exacto o header que empieza con el alias.

    Plantillas largas ('Cédula (Opcional si es menor…)') matchean 'cedula'.
    """
    mapa = {key(k): v for k, v in row.items()}
    for alias in aliases:
        ka = key(alias)
        valor = mapa.get(ka)
        if valor not in (None, ""):
            return valor
    for alias in aliases:
        ka = key(alias)
        if not ka:
            continue
        for hk, v in mapa.items():
            if v in (None, ""):
                continue
            if hk.startswith(ka + " ") or hk.startswith(ka + "("):
                return v
    return ""


def texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def parse_bool(valor: Any) -> bool:
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        return valor != 0
    limpio = key(texto(valor))
    return limpio in {"si", "sí", "s", "true", "t", "1", "yes", "y", "positivo", "x"}


def inferir_flags_seguridad(obs: str) -> dict[str, bool]:
    """Deriva flags desde texto libre SIIPOL/contrainteligencia."""
    t = key(obs)
    return {
        "registro_policial": (
            "posee registros policiales" in t
            or "posee registro policial" in t
            or "registros policiales" in t
        ),
        # Filosofía operativa / SIIPOL: denuncia de persona extraviada
        # = búsqueda activa = solicitada (misma bandeja KPI Solicitados).
        "solicitado": (
            "se encuentra solicitado" in t
            or "solicitado por" in t
            or "persona extraviada" in t
            or "registra como persona extraviada" in t
            or "extraviada" in t
            or "extraviado" in t
        ),
        "deportado": "deportado" in t,
    }


def scrub_dato_politico(obs: str) -> str:
    """Quita firmas de referéndum y afiliación política; conserva policial."""
    if not obs:
        return ""
    limpio = re.sub(
        r"(?i)(,\s*)?(y\s+)?(s[ií]\s+)?firm[oó]\s+contra\s+el\s+presidente"
        r"(\s+en\s+\d{4})?(\s+y\s+(\d{4}|firm[oó]\s+contra\s+el\s+presidente"
        r"(\s+en\s+\d{4})?))?",
        "",
        obs,
    )
    limpio = re.sub(r"(?i)(,\s*)?(data\s+)?vente\s+venezuela", "", limpio)
    limpio = re.sub(
        r"(?i)(,\s*)?militante(\s+de)?(\s+pj)?(\s+distrito\s+capital)?"
        r"(\s+vente\s+venezuela)?",
        "",
        limpio,
    )
    limpio = re.sub(r"[.,;]\s*[.,;]+", ".", limpio)
    limpio = re.sub(r"\s{2,}", " ", limpio)
    limpio = re.sub(r",(\S)", r", \1", limpio)
    return limpio.strip(" ,.;")


def parse_cedula(raw: Any) -> tuple[str, str]:
    """Extrae (tipo_doc, numero). Rechaza texto basura ('no posee', 'no tiene').

    Solo acepta patrones de documento real: V/E/P + dígitos, o solo dígitos
    (con separadores de miles). Núcleos familiares (12345678-1) → sin cédula.
    Nunca persiste texto libre en documento: si hay letras basura, se descartan
    o la fila queda sin cédula.
    """
    valor = texto(raw).upper().strip()
    if not valor or es_marcador_sin_cedula(valor):
        return "", ""

    # Núcleos familiares: menor sin cédula = doc adulto + sufijo corto (1-2 dígitos).
    if re.match(r"^([VEP][-\s.]?)?[\d.,]+-\d{1,2}$", valor):
        return "", ""

    # V/E/P + número (puntos/comas/espacios/guiones como separadores).
    match = re.match(r"^([VEP])[-\s.]?([\d.,\s\-]+)$", valor)
    if match:
        letra, bruto = match.group(1), match.group(2)
        num = re.sub(r"\D", "", bruto)
        if 5 <= len(num) <= 10 and not set(num) <= {"0"}:
            if int(num) > 70_000_000:
                return "E", num
            return letra, num
        return "", ""

    # Solo dígitos / separadores de miles — sin letras residuales.
    if re.fullmatch(r"[\d.,\s\-]+", valor):
        num = re.sub(r"\D", "", valor)
        if 5 <= len(num) <= 10 and not set(num) <= {"0"}:
            if int(num) > 70_000_000:
                return "E", num
            return "V", num
        return "", ""

    # Mezcla texto+número: conservar solo dígitos; nunca persistir el texto.
    digitos = re.sub(r"\D", "", valor)
    if 5 <= len(digitos) <= 10 and not set(digitos) <= {"0"}:
        if valor.lstrip().startswith("E"):
            letra = "E"
        elif valor.lstrip().startswith("P"):
            letra = "P"
        elif valor.lstrip().startswith("V"):
            letra = "V"
        elif int(digitos) > 70_000_000:
            letra = "E"
        else:
            letra = "V"
        return letra, digitos

    # Texto libre sin número usable: sin cédula.
    return "", ""


PARTICULAS_NOMBRE = frozenset(
    {"de", "del", "la", "las", "los", "y", "san", "santa", "da", "do", "das", "dos", "e"}
)


def _tokens_persona(texto_raw: str) -> list[str]:
    """Tokens de nombre/apellido: colapsa espacios y quita puntuación residual."""
    limpio = re.sub(r"\s+", " ", texto(texto_raw)).strip()
    if not limpio:
        return []
    out: list[str] = []
    for t in limpio.split():
        tok = re.sub(r"^[^\wÁÉÍÓÚÜÑáéíóúüñ]+|[^\wÁÉÍÓÚÜÑáéíóúüñ]+$", "", t)
        if tok:
            out.append(tok)
    return out


def normalizar_caso_persona(valor: str) -> str:
    """Title Case con partículas españolas en minúscula (De La Rosa → de la Rosa no; La Rosa)."""
    partes = _tokens_persona(valor)
    if not partes:
        return ""
    out: list[str] = []
    for i, tok in enumerate(partes):
        low = strip_accents(tok).lower()
        if i > 0 and low in PARTICULAS_NOMBRE:
            out.append(low)
        elif len(tok) == 1 and tok.isalpha():
            # Inicial suelta (B, J) → mayúscula simple
            out.append(tok.upper())
        else:
            out.append(tok[:1].upper() + tok[1:].lower() if tok else "")
    return " ".join(p for p in out if p)


def split_nombre(nombre: str) -> tuple[str, str]:
    partes = _tokens_persona(nombre)
    if not partes:
        return "", ""
    if len(partes) == 1:
        return partes[0], ""
    return partes[0], " ".join(partes[1:])


def split_apellido(apellido: str) -> tuple[str, str]:
    """Separa apellidos; une compuestos con partícula (La Rosa, De Los Santos)."""
    partes = _tokens_persona(apellido)
    if not partes:
        return "", ""
    if len(partes) == 1:
        return partes[0], ""
    # Primer apellido compuesto: La Rosa / De La Cruz / San Juan
    i = 0
    if strip_accents(partes[0]).lower() in PARTICULAS_NOMBRE:
        i = 1
        while i < len(partes) - 1 and strip_accents(partes[i]).lower() in PARTICULAS_NOMBRE:
            i += 1
        if i < len(partes):
            i += 1
        return " ".join(partes[:i]), " ".join(partes[i:])
    return partes[0], " ".join(partes[1:])


def split_nombre_completo(nombre: str) -> tuple[str, str, str, str]:
    """Separa nombre completo cuando planilla no trae columnas individuales."""
    partes = _tokens_persona(nombre)
    if len(partes) < 2:
        return (partes[0] if partes else ""), "", "", ""
    if len(partes) == 2:
        return partes[0], "", partes[1], ""
    if len(partes) == 3:
        return partes[0], partes[1], partes[2], ""
    return partes[0], partes[1], partes[2], " ".join(partes[3:])


# Apellidos vistos en la planilla (rellenado en main) para despegar ALLCAPS.
APELLIDOS_CONOCIDOS: set[str] = set()


def despegar_nombre_pegado(nombre: str) -> str:
    """Separa PascalCase/CamelCase sin espacios (ErinderAlexdick → Erinder Alexdick)."""
    t = texto(nombre)
    if not t or " " in t:
        return t
    if re.search(r"[a-z]", t) and re.search(r"[A-Z]", t):
        t = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", t)
        t = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", t)
    return t


def cargar_apellidos_conocidos(filas: list[dict[str, Any]]) -> None:
    """Indexa apellidos de filas con espacios para despegar nombres ALLCAPS pegados."""
    global APELLIDOS_CONOCIDOS
    apes: set[str] = set()
    for row in filas:
        for alias in (
            "nombre completo",
            "nombres y apellidos",
            "nombre y apellidos",
            "apellidos",
            "apellido",
            "primer apellido",
            "segundo apellido",
        ):
            val = texto(pick(row, alias))
            if not val:
                continue
            if "apellido" in alias and " " not in val and len(val) >= 3:
                apes.add(strip_accents(val).upper())
                continue
            partes = val.split()
            if len(partes) < 2:
                continue
            # Solo últimos 1–2 tokens = apellidos (evita indexar nombres de pila).
            for tok in partes[-2:]:
                tok_u = strip_accents(tok).upper()
                if len(tok_u) >= 3 and tok_u.isalpha():
                    apes.add(tok_u)
    # Semilla mínima frecuente en planillas VE
    apes.update(
        {
            "GARCIA", "GONZALEZ", "RODRIGUEZ", "MARTINEZ", "HERNANDEZ", "LOPEZ",
            "PEREZ", "SANCHEZ", "RAMIREZ", "TORRES", "FLORES", "RIVERA", "GOMEZ",
            "DIAZ", "MORALES", "REYES", "CRUZ", "ORTIZ", "GUTIERREZ", "CASTILLO",
            "ROJAS", "MOYA", "MATHEUS", "QUIROZ", "VERA", "SALAZAR", "ARAUJO",
            "HIPOLITO", "PALOMINO", "PARACO", "ANDRADE", "SALVAREZ", "ALVAREZ",
            "PACHECO", "ZAMBRANO", "ZAMBRZO", "CORPA", "SALVARADO", "ALVARADO",
            "SUAREZ", "GUZMAN", "MUJICA", "SANTANA", "SALAVARRIA", "SOJO",
            "DELGADO",
        }
    )
    APELLIDOS_CONOCIDOS = apes


def separar_allcaps_pegado(nombre: str) -> tuple[str, str, str, str] | None:
    """Si nombre es UN token ALLCAPS, pela hasta 2 apellidos conocidos del final."""
    t = texto(nombre)
    if not t or " " in t or not APELLIDOS_CONOCIDOS:
        return None
    if not re.fullmatch(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+", t):
        return None
    # Solo si no hay minúsculas intercaladas (ya cubierto por CamelCase)
    if re.search(r"[a-z]", t) and re.search(r"[A-Z]", t):
        return None
    rest = strip_accents(t).upper()
    apellidos_ord = sorted(APELLIDOS_CONOCIDOS, key=len, reverse=True)
    found: list[str] = []
    for _ in range(2):
        hit = next(
            (a for a in apellidos_ord if rest.endswith(a) and len(rest) >= len(a) + 3),
            None,
        )
        if not hit:
            break
        found.insert(0, hit.title())
        rest = rest[: -len(hit)]
    if not found or len(rest) < 2:
        return None
    return rest.title(), "", found[0], (found[1] if len(found) > 1 else "")


def parse_nucleo_familiar(raw: Any) -> tuple[str, str, str] | None:
    """Detecta cédula de responsable + sufijo de menor (17089732-1).

    Retorna (tipo_doc_jefe, documento_jefe, sufijo) o None.
    """
    valor = texto(raw).upper().strip()
    if not valor:
        return None
    match = re.match(r"^([VEP])?[-\s.]?([\d.,\s]+)-(\d{1,2})$", valor)
    if not match:
        return None
    letra = match.group(1) or ""
    num = re.sub(r"\D", "", match.group(2))
    sufijo = match.group(3)
    if not (5 <= len(num) <= 10) or set(num) <= {"0"}:
        return None
    if not letra:
        letra = "E" if int(num) > 70_000_000 else "V"
    return letra, num, sufijo


def mapear_parentesco_jefe(raw: str) -> str:
    p = key(raw)
    if not p:
        return "Hijo/a"
    if "hijo" in p or "hija" in p:
        return "Hijo/a"
    if "niet" in p:
        return "Nieto/a"
    if "sobrin" in p:
        return "Sobrino/a"
    if "herman" in p:
        return "Hermano/a"
    if "padre" in p or "madre" in p or "papa" in p or "mama" in p:
        return "Padre/Madre"
    if "espos" in p or "conyug" in p or "mujer" in p or "marido" in p:
        return "Cónyuge"
    if "abuel" in p:
        return "Abuelo/a"
    if "tio" in p or "tia" in p:
        return "Tío/a"
    return "Otro familiar"


def normalizar_sexo(valor: Any) -> str:
    limpio = key(texto(valor))
    if limpio in {"m", "masculino", "h", "hombre"}:
        return "M"
    if limpio in {"f", "femenino", "femenina", "mujer"}:
        return "F"
    return ""


def parse_edad(valor: Any) -> int | None:
    raw = texto(valor)
    if not raw:
        return None
    match = re.search(r"\d+", raw)
    if not match:
        return None
    edad = int(match.group(0))
    return edad if 0 <= edad <= 120 else None


def edad_nexus(body: dict[str, Any]) -> int | None:
    edad = body.get("edad")
    return int(edad) if isinstance(edad, int) and 0 <= edad <= 120 else None


def aplicar_nexus(payload: dict[str, Any], body: dict[str, Any], fuente: str = "nexus") -> None:
    payload["primer_nombre"] = texto(body.get("primer_nombre")) or payload["primer_nombre"]
    payload["segundo_nombre"] = texto(body.get("segundo_nombre")) or payload["segundo_nombre"]
    payload["primer_apellido"] = texto(body.get("primer_apellido")) or payload["primer_apellido"]
    payload["segundo_apellido"] = texto(body.get("segundo_apellido")) or payload["segundo_apellido"]
    payload["edad"] = "" if edad_nexus(body) is None else str(edad_nexus(body))
    sexo = normalizar_sexo(body.get("sexo"))
    if sexo:
        payload["sexo"] = sexo
    telefonos = body.get("telefonos")
    if not payload.get("telefono") and isinstance(telefonos, list) and telefonos:
        payload["telefono"] = texto(telefonos[0])
    payload["verificado_nexus"] = True
    payload["verificado_nexus_fuente"] = fuente if fuente in {"nexus", "cache"} else "nexus"


# Respuestas SIIPOL que confirman verificación pero no aportan obs útil.
_SIIPOL_MARCA_SIN_OBS = frozenset(
    {
        "no registra",
        "menor",
        "si",
        "sí",
        "no",
        "true",
        "false",
        "verificado",
        "ok",
        "n/a",
        "na",
        "-",
    }
)


def fila_a_payload(
    row: dict[str, Any],
    centros: list[CentroApp],
    centro_forzado: str | None,
    col_centro: str | None,
    *,
    ignorar_centro: bool = False,
) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    ced_raw = pick(row, "cedula", "cédula", "cedúla", "documento", "ci", "doc")
    nucleo = parse_nucleo_familiar(ced_raw)
    tipo_doc, documento = parse_cedula(ced_raw)
    tipo_doc_col = texto(pick(row, "tipo doc.", "tipo doc", "tipo_doc", "tipo documento")).upper()
    if tipo_doc_col in {"V", "E", "P"} and documento:
        tipo_doc = tipo_doc_col

    # Núcleo familiar (doc-responsable-N): menor sin cédula propia.
    jefe_tipo_doc = ""
    jefe_documento = ""
    parentesco_jefe = ""
    if nucleo:
        tipo_doc, documento = "", ""
        jefe_tipo_doc, jefe_documento, _sufijo = nucleo
        parentesco_jefe = mapear_parentesco_jefe(
            texto(pick(row, "parentesco", "parentesco jefe", "parentesco_jefe", "relacion", "relación"))
        )

    # Preferir columna única primero: pick("nombre") matchea "nombre completo" por prefijo.
    completo = texto(
        pick(
            row,
            "nombre completo",
            "nombre_completo",
            "nombre_apellido",
            "nombre apellido",
            "nombres_apellidos",
            "nombre y apellido",
            "nombres y apellidos",
            "nombre y apellidos",
            "habitante nombres y apellidos",
            "habitante (nombres y apellidos)",
            "beneficiario",
            "persona",
        )
    )
    nombre = texto(pick(row, "nombres", "nombre", "primer nombre", "primer_nombre"))
    segundo_nombre_col = texto(pick(row, "segundo nombre", "segundo_nombre"))
    apellido = texto(pick(row, "apellidos", "apellido", "primer apellido", "primer_apellido"))
    segundo_apellido_col = texto(pick(row, "segundo apellido", "segundo_apellido"))

    # Planillas mixtas: nombre vacío y nombre completo volcado en APELLIDO.
    if not completo and not nombre and apellido and len(_tokens_persona(apellido)) >= 2:
        completo = apellido
        apellido = ""

    completo = despegar_nombre_pegado(completo)
    nombre = despegar_nombre_pegado(nombre)
    apellido = despegar_nombre_pegado(apellido)

    allcaps = separar_allcaps_pegado(completo) or separar_allcaps_pegado(nombre)
    if allcaps:
        primer_nombre, segundo_nombre, primer_apellido, segundo_apellido = allcaps
    elif completo:
        primer_nombre, segundo_nombre, primer_apellido, segundo_apellido = split_nombre_completo(completo)
    else:
        primer_nombre, resto_nombre = split_nombre(nombre)
        primer_apellido, resto_apellido = split_apellido(apellido)
        segundo_nombre = segundo_nombre_col or resto_nombre
        segundo_apellido = segundo_apellido_col or resto_apellido

    primer_nombre = normalizar_caso_persona(primer_nombre)
    segundo_nombre = normalizar_caso_persona(segundo_nombre)
    primer_apellido = normalizar_caso_persona(primer_apellido)
    segundo_apellido = normalizar_caso_persona(segundo_apellido)

    nombre_centro = texto(row.get(col_centro, "")) if col_centro else ""
    if not nombre_centro:
        nombre_centro = texto(pick(row, "campamento", "centro", "escuela", "refugio", "institucion", "institución"))
    if ignorar_centro and not centro_forzado:
        centro_id, centro_raw, match = "", "", "omitido"
    else:
        centro_id, centro_raw, match = resolver_centro(nombre_centro, centros, centro_forzado)
        if not centro_id:
            return None, {
                "error": "centro_inactivo" if match == "inactivo" else "sin_centro",
                "nombre_centro_raw": centro_raw or nombre_centro,
            }

    edad = parse_edad(pick(row, "edad", "age"))
    tipo_registro = texto(pick(row, "tipo de registro", "tipo registro", "tipo registro policial"))
    descripcion_verificacion = texto(
        pick(
            row,
            "descripcion verificacion",
            "descripción verificación",
            "descripcion (verificacion)",
            "descripción (verificación)",
        )
    )
    # Columna "Verificado SIIPOL" del consolidado: NO REGISTRA / delito / firma / MENOR.
    col_verificado_siipol = texto(
        pick(
            row,
            "verificado siipol",
            "verificacion siipol",
            "verificación siipol",
            "estado siipol",
            "resultado siipol",
        )
    )
    observaciones_seguridad_col = texto(
        pick(
            row,
            "observaciones seguridad",
            "observacion seguridad",
            "observación seguridad",
        )
    )
    observaciones_generales = texto(
        pick(
            row,
            "observaciones",
            "observacion",
            "observación",
            "informacion de interes",
            "información de interés",
            "sistemas de contrainteligencia y siipol",
            "sistemas de contrainteligencia",
        )
    )
    obs_desde_verificado = ""
    if col_verificado_siipol and key(col_verificado_siipol) not in _SIIPOL_MARCA_SIN_OBS:
        obs_desde_verificado = col_verificado_siipol
    observaciones = (
        observaciones_seguridad_col
        or descripcion_verificacion
        or obs_desde_verificado
        or observaciones_generales
    )
    flags_texto = inferir_flags_seguridad(observaciones)
    registro_policial = parse_bool(
        pick(
            row,
            "tiene registro policial",
            "registro policial",
            "reg policial",
            "reg. policial",
            "con reg policial",
            "con registro policial",
        )
    ) or flags_texto["registro_policial"]
    solicitado = parse_bool(
        pick(row, "está solicitado", "esta solicitado", "solicitado", "requerido")
    ) or flags_texto["solicitado"]
    deportado = parse_bool(pick(row, "deportado")) or flags_texto["deportado"]
    # Dato político sensible: nunca persistir firma de referéndum ni afiliación.
    # Evidencia SIIPOL se mide ANTES del scrub: "NO REGISTRA" / "SIN INFORMACIÓN" también cuentan.
    evidencia_siipol = (
        bool(col_verificado_siipol)
        or bool(descripcion_verificacion)
        or bool(observaciones_seguridad_col)
    )
    observaciones = scrub_dato_politico(observaciones)
    verificado_siipol = evidencia_siipol or any(
        (registro_policial, solicitado, deportado, bool(tipo_registro))
    )

    payload: dict[str, Any] = {
        "centro_id": centro_id,
        "nombre_centro_raw": centro_raw or nombre_centro,
        "centro_match": match,
        "primer_nombre": primer_nombre,
        "segundo_nombre": segundo_nombre,
        "primer_apellido": primer_apellido,
        "segundo_apellido": segundo_apellido,
        "edad": "" if edad is None else str(edad),
        "tipo_doc": tipo_doc,
        "documento": documento,
        "jefe_tipo_doc": jefe_tipo_doc,
        "jefe_documento": jefe_documento,
        "parentesco_jefe": parentesco_jefe,
        "sexo": normalizar_sexo(pick(row, "sexo", "genero", "género")),
        "telefono": texto(
            pick(
                row,
                "telefono",
                "teléfono",
                "telefono principal",
                "teléfono principal",
                "celular",
                "phone",
            )
        ),
        "embarazada": parse_bool(pick(row, "embarazada", "embarazo")),
        "discapacidad": parse_bool(pick(row, "discapacidad", "discapacitado")),
        "discapacidad_detalle": texto(pick(row, "discapacidad detalle", "detalle discapacidad")),
        "enfermedad": parse_bool(pick(row, "enfermedad", "patologia", "patología")),
        "enfermedad_detalle": texto(pick(row, "enfermedad detalle", "detalle enfermedad", "patologia detalle")),
        "pais": texto(pick(row, "pais", "país")) or "Venezuela",
        "estado_federativo": texto(pick(row, "estado", "estado_federativo")),
        "municipio": texto(pick(row, "municipio")),
        "parroquia": texto(pick(row, "parroquia")),
        "calle": texto(pick(row, "direccion", "dirección", "calle", "sector")),
        "casa_edificio": texto(
            pick(
                row,
                "casa",
                "edificio",
                "casa_edificio",
                "aula",
                "ubicacion bloque carpa",
                "ubicacion (bloque/carpa)",
                "ubicacion",
                "ubicación",
            )
        ),
        "registro_policial": registro_policial,
        "solicitado": solicitado,
        "firmo_contra_presidente": False,
        "deportado": deportado,
        "tipo_registro_policial": tipo_registro,
        "observaciones_seguridad": observaciones,
        "verificado_siipol": verificado_siipol,
        "verificado_nexus": False,
        "verificado_nexus_fuente": "",
    }
    return payload, None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--archivo", required=True, type=Path)
    ap.add_argument("--centro-id", default=None, help="Fuerza centro_id si el Excel es de un solo campamento")
    ap.add_argument("--col-centro", default=None, help="Columna con nombre del campamento")
    ap.add_argument("--con-nexus", action="store_true", help="Consulta Nexus por cada cédula V/E y prioriza identidad oficial")
    ap.add_argument(
        "--solo-cache-nexus",
        action="store_true",
        help="Con --con-nexus, reutiliza caché existente sin consultar cédulas pendientes",
    )
    ap.add_argument(
        "--concurrency",
        type=int,
        default=DEFAULT_CONCURRENCY,
        help=f"Consultas Nexus simultáneas y por segundo (default calibrado: {DEFAULT_CONCURRENCY})",
    )
    ap.add_argument(
        "--rate",
        type=float,
        default=DEFAULT_RATE,
        help=f"Segundos entre consultas si --concurrency=1 (default {DEFAULT_RATE})",
    )
    ap.add_argument(
        "--circuit-breaker",
        type=int,
        default=DEFAULT_CIRCUIT_BREAKER,
        help=f"Corta tras N fallos consecutivos de infraestructura (default {DEFAULT_CIRCUIT_BREAKER})",
    )
    ap.add_argument(
        "--timeout-nexus",
        type=float,
        default=DEFAULT_NEXUS_TIMEOUT,
        help=f"Timeout por consulta Nexus en segundos (default {DEFAULT_NEXUS_TIMEOUT:g})",
    )
    ap.add_argument("--aplicar", action="store_true", help="Escribe en BD vía censo_importar_lote")
    ap.add_argument(
        "--solo-marcar-siipol",
        action="store_true",
        help="No reimporta filas; solo aplica evidencia SIIPOL sobre registros existentes",
    )
    ap.add_argument(
        "--reconciliar-siipol",
        action="store_true",
        help="Usa Documento como lista autoritativa SIIPOL; no importa personas",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="No importa censo; las verificaciones Nexus nuevas sí quedan en nexus_consultas",
    )
    ap.add_argument("--lote", type=int, default=200)
    ap.add_argument("--solo-con-cedula", action="store_true", help="Omite filas sin documento")
    ap.add_argument(
        "--permitir-omisiones",
        action="store_true",
        help="Permite aplicar aunque existan filas inválidas o campamentos sin resolver",
    )
    ap.add_argument(
        "--omitir-firmo-presidente",
        action="store_true",
        help=argparse.SUPPRESS,  # legado: siempre omitido; flag no-op
    )
    ap.add_argument("--json-out", type=Path, default=None)
    args = ap.parse_args()

    if not args.archivo.exists():
        raise SystemExit(f"No existe {args.archivo}")
    if not 1 <= args.concurrency <= 10:
        raise SystemExit("--concurrency debe estar entre 1 y 10")
    if args.rate < 0:
        raise SystemExit("--rate no puede ser negativo")
    if args.circuit_breaker < 1:
        raise SystemExit("--circuit-breaker debe ser al menos 1")
    if args.timeout_nexus <= 0:
        raise SystemExit("--timeout-nexus debe ser mayor que 0")
    if args.solo_cache_nexus and not args.con_nexus:
        raise SystemExit("--solo-cache-nexus requiere --con-nexus")
    if args.reconciliar_siipol and args.solo_marcar_siipol:
        raise SystemExit("--reconciliar-siipol y --solo-marcar-siipol son excluyentes")

    dry = args.dry_run or not args.aplicar
    print(f"Excel: leyendo {args.archivo.name}…", file=sys.stderr, flush=True)
    rows, hoja_datos, columnas = leer_filas(args.archivo)
    cargar_apellidos_conocidos(rows)
    print(
        f"Excel: hoja «{hoja_datos}» · {len(rows)} filas · {len(columnas)} columnas",
        file=sys.stderr,
        flush=True,
    )
    print("Supabase: autenticando…", file=sys.stderr, flush=True)
    url, anon, gateway = cargar_env()
    jwt = autenticar(url, anon)

    if args.reconciliar_siipol:
        documentos_filas = [
            parse_cedula(pick(row, "cedula", "cédula", "documento", "ci", "doc"))[1]
            for row in rows
        ]
        documentos = sorted({documento for documento in documentos_filas if documento})
        reporte_siipol = {
            "archivo": args.archivo.name,
            "hoja": hoja_datos,
            "filas_leidas": len(rows),
            "documentos_validos": sum(bool(documento) for documento in documentos_filas),
            "documentos_unicos": len(documentos),
            "documentos_duplicados": sum(
                cantidad - 1
                for documento, cantidad in Counter(documentos_filas).items()
                if documento and cantidad > 1
            ),
            "sin_documento": sum(not documento for documento in documentos_filas),
        }
        print(json.dumps(reporte_siipol, ensure_ascii=False, indent=2))
        if dry:
            print("Dry-run: no se modificaron marcas SIIPOL.", file=sys.stderr)
            return 0
        resultado = rpc(
            url,
            anon,
            jwt,
            "censo_reconciliar_siipol",
            {"p_documentos": documentos, "p_fuente": args.archivo.name},
        )
        print(json.dumps(resultado, ensure_ascii=False, indent=2))
        return 0

    print("Supabase: cargando campamentos…", file=sys.stderr, flush=True)
    centros = listar_centros(url, anon, jwt)

    preparadas: list[dict[str, Any]] = []
    ok: list[dict[str, Any]] = []
    errores: list[dict[str, Any]] = []
    nexus_errores: list[dict[str, Any]] = []
    conteos = {
        "con_cedula": 0,
        "sin_cedula": 0,
        "nexus_cache": 0,
        "nexus_ya_verificadas_unicas": 0,
        "nexus_consultadas": 0,
        "nexus_verificadas_nuevas": 0,
        "nexus_ok": 0,
        "nexus_error": 0,
        "nexus_omitidas_circuit_breaker": 0,
        "nexus_omitidas_solo_cache": 0,
        "nexus_omitidas_doc_invalido": 0,
        "solicitados": 0,
        "registro_policial": 0,
        "verificados_siipol": 0,
    }

    for row in rows:
        _, documento_entrada = parse_cedula(pick(row, "cedula", "cédula", "documento", "ci", "doc"))
        if documento_entrada:
            conteos["con_cedula"] += 1
        else:
            conteos["sin_cedula"] += 1
        payload, error = fila_a_payload(
            row,
            centros,
            args.centro_id,
            args.col_centro,
            ignorar_centro=args.solo_marcar_siipol,
        )
        if error:
            errores.append(error)
            continue
        if payload is None:
            errores.append({"error": "fila_incompleta", "row": row})
            continue

        if not payload.get("documento") and args.solo_con_cedula:
            continue
        preparadas.append(payload)

    cache_persistente: dict[tuple[str, str], dict[str, Any]] = {}
    fichas_nuevas: dict[tuple[str, str], dict[str, Any]] = {}
    candidatos: set[tuple[str, str]] = set()
    if args.con_nexus:
        omitidas_doc_invalido = 0
        for payload in preparadas:
            tipo_doc = texto(payload.get("tipo_doc")).upper()
            documento = texto(payload.get("documento"))
            if tipo_doc not in LETRAS_NEXUS or not documento:
                continue
            # V/E consultable: 6–8 dígitos; basura (9+, 00000, etc.) no va a Nexus.
            if len(documento) < 6 or len(documento) > 8 or set(documento) <= {"0"}:
                omitidas_doc_invalido += 1
                continue
            candidatos.add((tipo_doc, documento))
        if omitidas_doc_invalido:
            conteos["nexus_omitidas_doc_invalido"] = omitidas_doc_invalido
            print(
                f"Nexus: {omitidas_doc_invalido} cédulas con formato inválido omitidas",
                file=sys.stderr,
                flush=True,
            )

        print("Nexus: cargando verificaciones existentes…", file=sys.stderr, flush=True)
        cache_persistente = cargar_cache_nexus(url, anon, jwt)
        pendientes = sorted(candidatos - cache_persistente.keys())
        conteos["nexus_ya_verificadas_unicas"] = len(candidatos) - len(pendientes)
        pendientes_totales = len(pendientes)
        print(
            f"Nexus: {len(candidatos)} cédulas únicas · "
            f"{len(candidatos) - len(pendientes)} ya verificadas · "
            f"{len(pendientes)} pendientes · concurrencia {args.concurrency}",
            file=sys.stderr,
        )
        if args.solo_cache_nexus:
            conteos["nexus_omitidas_solo_cache"] = pendientes_totales
            pendientes = []
            print(
                f"Nexus: modo solo caché · {pendientes_totales} pendientes usarán datos del Excel",
                file=sys.stderr,
                flush=True,
            )

        if pendientes:
            jwt_ref = {"token": jwt}

            def flush_incremental(lote: dict[tuple[str, str], dict[str, Any]]) -> None:
                jwt_ref["token"] = guardar_cache_nexus(url, anon, jwt_ref["token"], lote)

            fichas_nuevas, nexus_errores, omitidas = consultar_nexus_concurrente(
                gateway,
                jwt,
                pendientes,
                args.concurrency,
                args.rate,
                args.circuit_breaker,
                args.timeout_nexus,
                flush_fichas=flush_incremental,
            )
            jwt = jwt_ref["token"]
            conteos["nexus_consultadas"] = len(fichas_nuevas) + len(nexus_errores)
            conteos["nexus_verificadas_nuevas"] = len(fichas_nuevas)
            conteos["nexus_omitidas_circuit_breaker"] = omitidas
            # Flush incremental ya persistió; noop por si quedó cola.
            jwt = guardar_cache_nexus(url, anon, jwt, {})

        fichas_disponibles = {**cache_persistente, **fichas_nuevas}
        for payload in preparadas:
            clave = (
                texto(payload.get("tipo_doc")).upper(),
                texto(payload.get("documento")),
            )
            ficha = fichas_disponibles.get(clave)
            if ficha is not None:
                fuente_nexus = "cache" if clave in cache_persistente else "nexus"
                aplicar_nexus(payload, ficha, fuente_nexus)
                conteos["nexus_ok"] += 1
                if clave in cache_persistente:
                    conteos["nexus_cache"] += 1
            elif clave[0] in LETRAS_NEXUS and clave[1]:
                conteos["nexus_error"] += 1

    for payload in preparadas:
        if not payload.get("primer_nombre") or not payload.get("primer_apellido"):
            errores.append({"error": "sin_nombre", "documento": payload.get("documento")})
            continue

        payload["firmo_contra_presidente"] = False
        payload["observaciones_seguridad"] = scrub_dato_politico(
            texto(payload.get("observaciones_seguridad"))
        )

        if payload.get("solicitado"):
            conteos["solicitados"] += 1
        if payload.get("registro_policial"):
            conteos["registro_policial"] += 1
        if payload.get("verificado_siipol"):
            conteos["verificados_siipol"] += 1
        ok.append(payload)

    ok.sort(key=lambda f: (0 if f.get("documento") else 1, f.get("primer_apellido") or ""))
    documentos = [texto(f.get("documento")) for f in ok if f.get("documento")]
    documentos_repetidos = sum(cantidad - 1 for cantidad in Counter(documentos).values() if cantidad > 1)
    errores_por_tipo = Counter(texto(error.get("error")) or "desconocido" for error in errores)
    centros_con_error = Counter(
        texto(error.get("nombre_centro_raw"))
        for error in errores
        if error.get("nombre_centro_raw")
    )
    columnas_politicas = {
        "milita oposicion",
        "firmo contra presidente",
        "firmó contra presidente",
        "firmo vs pres",
        "firmo vs presidente",
        "firmo contra el gob",
        "firmó contra el gob.",
    }
    sensibles_ignoradas = [
        columna for columna in columnas if key(columna) in columnas_politicas
    ]
    reporte = {
        "archivo": args.archivo.name,
        "hoja": hoja_datos,
        "filas_leidas": len(rows),
        "listas": len(ok),
        **conteos,
        "documentos_repetidos": documentos_repetidos,
        "errores_match": len(errores),
        "errores_por_tipo": dict(sorted(errores_por_tipo.items())),
        "centros_con_error": dict(sorted(centros_con_error.items())),
        "columnas_sensibles_ignoradas": sensibles_ignoradas,
        "nexus_errores_muestra": [
            {"tipo_doc": error.get("tipo_doc"), "status": error.get("status")}
            for error in nexus_errores[:10]
        ],
    }
    print(json.dumps(reporte, ensure_ascii=False, indent=2))

    if args.json_out:
        args.json_out.write_text(
            json.dumps({"filas": ok, "errores": errores, "nexus_errores": nexus_errores}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"JSON escrito: {args.json_out}", file=sys.stderr)

    if dry:
        detalle_nexus = (
            " Verificaciones Nexus nuevas quedaron guardadas en nexus_consultas; "
            "--aplicar las reutilizará."
            if args.con_nexus and not args.solo_cache_nexus
            else ""
        )
        print(f"Dry-run: no se importó el censo.{detalle_nexus}", file=sys.stderr)
        return 0
    if errores and not args.permitir_omisiones:
        raise SystemExit(
            f"Importación cancelada: {len(errores)} filas no están listas. "
            "Corrija los errores o use --permitir-omisiones tras aprobar una importación parcial."
        )

    insertados = actualizados = omitidos = marcados_siipol = 0
    errores_rpc: list[Any] = []
    for i in range(0, len(ok), args.lote):
        chunk = ok[i : i + args.lote]
        result: object = {"modo": "solo_marcar_siipol"}
        if not args.solo_marcar_siipol:
            result = rpc(
                url,
                anon,
                jwt,
                "censo_importar_lote",
                {"p_filas": chunk, "p_meta": {"fuente_archivo": args.archivo.name}},
            )
            if isinstance(result, dict):
                insertados += int(result.get("insertados") or 0)
                actualizados += int(result.get("actualizados") or 0)
                omitidos += int(result.get("omitidos") or 0)
                err = result.get("errores") or []
                if isinstance(err, list):
                    errores_rpc.extend(err)
        if args.solo_marcar_siipol:
            resultado_siipol = rpc(
                url,
                anon,
                jwt,
                "censo_marcar_siipol_lote",
                {"p_filas": chunk, "p_fuente": args.archivo.name},
            )
            result = resultado_siipol
            if isinstance(resultado_siipol, dict):
                marcados_siipol += int(resultado_siipol.get("marcados_siipol") or 0)
        print(f"Lote {i // args.lote + 1}: {result}", file=sys.stderr)

    print(
        json.dumps(
            {
                "insertados": insertados,
                "actualizados": actualizados,
                "omitidos": omitidos,
                "marcados_siipol": marcados_siipol,
                "errores_rpc": errores_rpc[:30],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
